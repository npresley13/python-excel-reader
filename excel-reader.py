from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re

wb = load_workbook(filename = 'excel-reader/ecommerce-role.xlsx')
ws = wb.active
words = []


for cell in ws['A']:
    cellValue = str(cell.value)
    arr = re.split(r"[-;,.\s<>=/:\"\'(){}_+\]\\\[?$#!@%^&*-]\s*",cellValue)
    words = words + arr

df = pd.value_counts(np.array(words))
df.to_csv('ecom-role.csv')

print('Index:', df.index)
print('Values:', df.values)






